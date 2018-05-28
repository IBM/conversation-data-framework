"""
Copyright 2018 IBM Corporation
Licensed under the Apache License, Version 2.0 (the "License");
you may not use this file except in compliance with the License.
You may obtain a copy of the License at

    http://www.apache.org/licenses/LICENSE-2.0

Unless required by applicable law or agreed to in writing, software
distributed under the License is distributed on an "AS IS" BASIS,
WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
See the License for the specific language governing permissions and
limitations under the License.
"""

import os, re
import unidecode
from openpyxl import load_workbook
from wawCommons import printf, eprintf, toIntentName
from zipfile import BadZipfile
import DialogData as Dialog
from DialogData import DialogData

'''
Created on Jan 12, 2018
@author: alukes
'''
NAME_POLICY = 'soft'


class XLSXHandler(object):
    """ Converts Excel speadsheet into T2C data structures. """


    def __init__(self):
        self._dataBlocks = []
        self._labelsMap = {}
        self._dialogData = DialogData()


    def getDataBlocks(self):
        """ Return map with GoTo labels as keys and target Dialog intents as values. 
            This map is global across all processed Excel source files. """
        return self._dataBlocks


    def getLabelsMap(self):
        """ Return map with GoTo labels as keys and target Dialog intents as values. 
            This map is global across all processed Excel source files. """
        return self._labelsMap


    def getDialogData(self):
        return self._dialogData


    def addLabel(self, label, intent):
        self._labelsMap[label] = intent


    def addDataBlock(self, dataBlock):
        self._dataBlocks.append(dataBlock)


    def parseXLSXIntoDataBlocks(self, filename):
        """ Read Excel spreadsheet in T2C format. Store the data as tuples (domain, prefix, intent, rawBlock) into private field. """

        printf('Processing xlsx file: %s\n', filename)
        if not os.path.exists(filename):
            eprintf('Error: File does not exist: %s\n', filename)
            return {}
    
        try:
            domainName = unicode(toIntentName(NAME_POLICY, None, os.path.splitext(os.path.split(filename)[1])[0]), 'utf-8')
            workbook = load_workbook(filename=filename, read_only=True)
        except (IOError, BadZipfile):
            eprintf('Error: File does not seem to be a valid Excel spreadsheet: %s\n', filename)
            return {}
    
        # Process all the tabs of the file
        for sheet in workbook.worksheets:
            printf(' Sheet: %s\n', sheet.title)
            prefix = unicode(sheet.title, 'utf-8')
            currentBlock = []

            # Separate all data blocks in the sheet, if the currentBlock starts with header, it is considered to be part of currentBlock
            for row in sheet.iter_rows(max_col=4):
                validRow = False
                for columnIndex in range (0, 4):
                    if row[columnIndex] and row[columnIndex].value:
                        validRow = True
                if row[0].value and row[0].value.startswith('//'):
                    validRow = False

                # If empty line or header, we store the previous currentBlock-if any
                if not validRow:
                    if currentBlock:
                        self.__createBlock(domainName, prefix, currentBlock)
                    currentBlock = []
                else:
                    currentBlock.append((row[0].value.strip() if row[0].value else None,
                                         row[1].value.strip() if row[1].value else None,
                                         row[2].value.strip() if row[2].value else None,
                                         row[3].value.strip() if row[3].value else None))
            if currentBlock:
                self.__createBlock(domainName, prefix, currentBlock)


    def convertBlocksToDialogData(self):
        """ Read all parsed raw blocks of data and handle it depending on the type of block. """
        for domain, prefix, intent, block in self._dataBlocks:
            if not block or not isinstance(block[0], tuple) or not block[0][0]:
                printf('Warning: First cell of the data block does not contain any data. (domain=%s, prefix=%s, intent=%s)\n', domain, prefix, intent)
                continue

            if self.__isConditionBlock(block[0][0]):
                self.__handleConditionBlock(intent, block, domain)
            else:
                self.__handleIntentBlock(intent, block, domain)


    def __createBlock(self, domain, prefix, block):
        if not block or not block[0][0]:
            printf('Warning: First cell of the data block does not contain any data. (domain=%s, prefix=%s)\n', domain, prefix)
            return

        # Check if there's a label
        label = None
        firstCell = block[0][0]
        if firstCell.startswith(u':') and len(block) > 1:
            label = firstCell[1:]
            if label in self._labelsMap:
                printf('Warning: Found a label that has already been assigned to an intent and will be overwritten. Label: %s\n', label)
            del block[0]
            firstCell = block[0][0]

        # If it's entity block, load the entity
        if firstCell.startswith(u'@'):
            self.__handleEntityBlock(block)
            return
        
        # Check the intent name
        conditionHasX = Dialog.X_PLACEHOLDER in firstCell
        intent = firstCell

        if self.__isConditionBlock(firstCell):
            if conditionHasX and block[1][0]:
                intent = re.sub(Dialog.X_PLACEHOLDER, block[1][0], firstCell)
        else:
            if firstCell.startswith(u'#'):
                intent = firstCell[1:] 
            else:
                # Create fully qualified intent name
                intent = toIntentName(NAME_POLICY, domain + '_' + prefix + '_' + intent)

        self._dialogData.getIntentData(intent, domain)
        self._dataBlocks.append((domain, prefix, intent, block))
        if label:
            self._labelsMap[label] = intent


    def __isConditionBlock(self, firstCell):
        return Dialog.X_PLACEHOLDER in firstCell or len(re.sub('[^#$@&|]', '', firstCell)) > 1

            
    def __handleConditionBlock(self, intent, block, domain):
        """ Read condition definition from current block and save it into Dialog data structure. 
            Replace all <x> placeholders on the first line with values from remaining lines. """
        conditionTemplate = block[0][0]
        conditionHasX = Dialog.X_PLACEHOLDER in conditionTemplate
        intentData = None

        if not conditionHasX:
            intentData = self._dialogData.getIntentData(intent, domain)

        for row in block[1:]:
            if row[0] and conditionHasX:
                intent = re.sub(Dialog.X_PLACEHOLDER, row[0], conditionTemplate)
                intentData = self._dialogData.getIntentData(intent, domain)
            intentData.addRawOutput(row[1:], self._labelsMap)


    def __handleEntityBlock(self, block):
        """ Read entity definition from current block and save it into Dialog data structure. """
        entityName = block[0][0][1:]  # From the first cell (index[0][0]), take value without '@' at the beginning
        for output in block[1:]:
            if output[0]:
                self._dialogData.getEntity(entityName).append(output[0])
                

    def __handleIntentBlock(self, intent, block, domain):
        """ Read intent definition from current block and save it into Dialog data structure. """
        # blockLength = len(block)
        startsWithHash = block[0][0].startswith(u'#')

        if not startsWithHash and not block[0][1]:
            eprintf('Warning: Wrong intent definition format for line starting with: %s\n', intent)
            return

        intentData = self._dialogData.getIntentData(intent, domain)

        for row in block:
            if row[0] and not row[0].startswith(u'#'):
                intentData.addIntentAlternative(row[0])  # Collect intent definition
            if row[1]:
                intentData.addRawOutput(row[1:], self._labelsMap)  # Collect text output

